import os
from PIL import Image
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows


def process_images_to_excel(folder_path, output_file='Тестирование.xlsx'):
    filenames = []
    widths = []
    heights = []
    total_pixels = []

    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)

        if os.path.isfile(file_path) and filename.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp', '.gif', '.tiff')):
            try:
                with Image.open(file_path) as img:
                    width, height = img.size
                    total = width * height

                    # Добавляем данные в списки
                    filenames.append(filename)
                    widths.append(width)
                    heights.append(height)
                    total_pixels.append(total)

            except Exception as e:
                print(f"Ошибка при обработке файла {filename}: {e}")

    df = pd.DataFrame({
        'Filename': filenames,
        'Width (px)': widths,
        'Height (px)': heights,
        'Total Pixels': total_pixels
    })

    additional_columns = [
        'Inidications (reference)', 'Series number (reference)', 'Model (reference)', 'Rate (reference)',
        'Indications', 'Series number', 'Model', 'Rate',
        'Indications Match', 'Series Match', 'Model Match', 'Rate Match', 'Overall Match'
    ]

    for col in additional_columns:
        df[col] = ''

    wb = Workbook()
    ws = wb.active
    ws.title = 'Image Data'
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    header_font = Font(bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    column_widths = {
        'A': 30,  # Filename
        'B': 12,  # Width (px)
        'C': 13,  # Height (px)
        'D': 15,  # Total Pixels
        'E': 22,  # Inidications (reference)
        'F': 22,  # Series number (reference)
        'G': 20,  # Model (reference)
        'H': 15,  # Rate (reference)
        'I': 12,  # Indications
        'J': 15,  # Series number
        'K': 10,  # Model
        'L': 8,  # Rate        'M': 18,  # Indications Match
        'N': 14,  # Series Match
        'O': 12,  # Model Match
        'P': 11,  # Rate Match
        'Q': 14  # Overall Match
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    wb.save(output_file)
    print(f"Файл {output_file} успешно создан!")
    print(f"Обработано {len(filenames)} изображений")


if __name__ == "__main__":
    folder_path = input("Введите путь к папке с изображениями: ")
    process_images_to_excel(folder_path)