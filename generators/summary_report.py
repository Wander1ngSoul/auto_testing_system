import logging
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

logger = logging.getLogger(__name__)


def create_summary_sheet(wb, report_data):
    try:
        if 'Итоговый отчет' in wb.sheetnames:
            ws = wb['Итоговый отчет']
            wb.remove(ws)

        ws = wb.create_sheet('Итоговый отчет')
        COLORS = {
            'primary': '4F81BD',
            'secondary': '9BBB59',
            'accent': 'F79646',
            'background': 'F2F2F2',
            'success': 'C6EFCE',
            'warning': 'FFEB9C',
            'error': 'FFC7CE',
            'header': 'DCE6F1',
            'border': 'BFBFBF',
        }
        title_font = Font(bold=True, size=16, color=COLORS['primary'], name='Calibri')
        header_font = Font(bold=True, size=12, color=COLORS['primary'], name='Calibri')
        bold_font = Font(bold=True, size=11, color='000000', name='Calibri')
        normal_font = Font(size=11, color='000000', name='Calibri')
        highlight_font = Font(bold=True, size=12, color=COLORS['accent'], name='Calibri')

        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        right_alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color=COLORS['border']),
            right=Side(style='thin', color=COLORS['border']),
            top=Side(style='thin', color=COLORS['border']),
            bottom=Side(style='thin', color=COLORS['border'])
        )

        thick_border = Border(
            left=Side(style='thick', color=COLORS['primary']),
            right=Side(style='thick', color=COLORS['primary']),
            top=Side(style='thick', color=COLORS['primary']),
            bottom=Side(style='thick', color=COLORS['primary'])
        )

        _fill_report_data(ws, report_data, COLORS, title_font, header_font, normal_font,
                          bold_font, highlight_font, center_alignment, left_alignment,
                          right_alignment, thin_border, thick_border)

        _adjust_column_widths(ws)
        _adjust_row_heights(ws)

        _apply_background(ws, COLORS['background'])

        logger.info("Создан красивый лист с итоговым отчетом")
        return True

    except Exception as e:
        logger.error(f"Ошибка создания листа с отчетом: {str(e)}")
        return False


def _fill_report_data(ws, report_data, colors, title_font, header_font, normal_font,
                      bold_font, highlight_font, center_alignment, left_alignment,
                      right_alignment, thin_border, thick_border):
    current_row = 1

    ws.merge_cells(f'A{current_row}:F{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "🎯 ИТОГОВЫЙ ОТЧЕТ ТЕСТИРОВАНИЯ"
    title_cell.font = title_font
    title_cell.alignment = center_alignment
    title_cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type="solid")
    title_cell.border = thick_border
    current_row += 1

    ws.row_dimensions[current_row].height = 5
    current_row += 1

    current_row = _create_info_block(ws, current_row, "📊 ОБЩАЯ ИНФОРМАЦИЯ",
                                     [("Всего тестов", report_data['accuracy']['total_tests'])],
                                     colors, header_font, bold_font, normal_font,
                                     left_alignment, thin_border)

    current_row += 1

    accuracy_data = [
        ("📈 Показания", report_data['accuracy']['indications']),
        ("🔢 Серийные номера", report_data['accuracy']['series']),
        ("💻 Модели", report_data['accuracy']['model']),
        ("💰 Тарифы", report_data['accuracy']['rate'])
    ]

    accuracy_rows = []
    for label, data in accuracy_data:
        accuracy_rows.append(
            (label, f"✓ {data['correct']}/{report_data['accuracy']['total_tests']} ({data['accuracy']:.1f}%)"))

    current_row = _create_info_block(ws, current_row, "🎯 ТОЧНОСТЬ РАСПОЗНАВАНИЯ",
                                     accuracy_rows, colors, header_font, bold_font,
                                     normal_font, left_alignment, thin_border)

    current_row += 1
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = "🏆 Общая точность:"
    ws[f'A{current_row}'].font = highlight_font
    ws[f'A{current_row}'].alignment = left_alignment
    ws[f'A{current_row}'].border = thin_border
    ws[f'A{current_row}'].fill = PatternFill(start_color=colors['success'], end_color=colors['success'],
                                             fill_type="solid")

    ws.merge_cells(f'C{current_row}:F{current_row}')
    accuracy_value = report_data['accuracy']['overall']['accuracy']
    accuracy_color = colors['success'] if accuracy_value >= 80 else colors['warning'] if accuracy_value >= 60 else \
    colors['error']
    ws[f'C{current_row}'] = f"{accuracy_value:.1f}%"
    ws[f'C{current_row}'].font = Font(bold=True, size=14, color=accuracy_color, name='Calibri')
    ws[f'C{current_row}'].alignment = center_alignment
    ws[f'C{current_row}'].border = thin_border
    ws[f'C{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    current_row += 1

    stats_data = [
        ("📁 Всего файлов", report_data['total_images']),
        ("✅ Успешно обработано", report_data['successfully_processed']),
        ("❌ Ошибок", report_data['errors']),
        ("⏭️ Пропущено", report_data['skipped']),
        ("🎯 Успешность обработки", f"{report_data['success_rate']:.2f}%"),
        ("⏱️ Общее время", f"{report_data['total_time_seconds']:.2f} сек"),
        ("⚡ Среднее время", f"{report_data['average_time_per_image']:.2f} сек/изобр"),
        ("🚀 Скорость", f"{report_data['images_per_minute']:.2f} изобр/мин")
    ]

    current_row = _create_info_block(ws, current_row, "📈 СТАТИСТИКА ОБРАБОТКИ",
                                     stats_data, colors, header_font, bold_font,
                                     normal_font, left_alignment, thin_border)

    current_row += 1
    ws.merge_cells(f'A{current_row}:B{current_row}')
    ws[f'A{current_row}'] = "🕒 Завершено:"
    ws[f'A{current_row}'].font = bold_font
    ws[f'A{current_row}'].alignment = left_alignment
    ws[f'A{current_row}'].border = thin_border
    ws[f'A{current_row}'].fill = PatternFill(start_color=colors['header'], end_color=colors['header'],
                                             fill_type="solid")

    ws.merge_cells(f'C{current_row}:F{current_row}')
    ws[f'C{current_row}'] = report_data['completion_time']
    ws[f'C{current_row}'].font = normal_font
    ws[f'C{current_row}'].alignment = left_alignment
    ws[f'C{current_row}'].border = thin_border
    ws[f'C{current_row}'].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    _add_outline_border(ws, 1, current_row, 1, 6, thick_border)


def _create_info_block(ws, start_row, title, data, colors, header_font, bold_font,
                       normal_font, alignment, border):
    ws.merge_cells(f'A{start_row}:F{start_row}')
    cell = ws[f'A{start_row}']
    cell.value = title
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border
    cell.fill = PatternFill(start_color=colors['header'], end_color=colors['header'], fill_type="solid")
    current_row = start_row + 1

    for i, (label, value) in enumerate(data):
        ws.merge_cells(f'A{current_row}:B{current_row}')
        ws[f'A{current_row}'] = label
        ws[f'A{current_row}'].font = bold_font
        ws[f'A{current_row}'].alignment = alignment
        ws[f'A{current_row}'].border = border

        fill_color = colors['background'] if i % 2 == 0 else "FFFFFF"
        ws[f'A{current_row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        ws.merge_cells(f'C{current_row}:F{current_row}')
        ws[f'C{current_row}'] = value
        ws[f'C{current_row}'].font = normal_font
        ws[f'C{current_row}'].alignment = alignment
        ws[f'C{current_row}'].border = border
        ws[f'C{current_row}'].fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

        current_row += 1

    return current_row


def _adjust_column_widths(ws):
    column_widths = {
        'A': 22, 'B': 3, 'C': 28, 'D': 3, 'E': 3, 'F': 3
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width


def _adjust_row_heights(ws):
    for row in range(1, ws.max_row + 1):
        if row == 1:  # Заголовок
            ws.row_dimensions[row].height = 30
        elif any(cell.value and '🚀' in str(cell.value) for cell in ws[row]):  # Заголовки блоков
            ws.row_dimensions[row].height = 25
        else:
            ws.row_dimensions[row].height = 20


def _apply_background(ws, color):
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if not cell.fill.start_color.index:  # Если ячейка без заливки
                cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def _add_outline_border(ws, start_row, end_row, start_col, end_col, border):
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            current_border = cell.border

            new_border = Border(
                left=border.left if col == start_col else current_border.left,
                right=border.right if col == end_col else current_border.right,
                top=border.top if row == start_row else current_border.top,
                bottom=border.bottom if row == end_row else current_border.bottom
            )
            cell.border = new_border

def extract_report_data_from_logs(log_text):
    lines = log_text.split('\n')
    report_data = {
        'accuracy': {
            'total_tests': 0,
            'indications': {'correct': 0, 'accuracy': 0},
            'series': {'correct': 0, 'accuracy': 0},
            'model': {'correct': 0, 'accuracy': 0},
            'rate': {'correct': 0, 'accuracy': 0},
            'overall': {'correct': 0, 'accuracy': 0}
        },
        'total_images': 0,
        'successfully_processed': 0,
        'errors': 0,
        'skipped': 0,
        'success_rate': 0,
        'total_time_seconds': 0,
        'average_time_per_image': 0,
        'images_per_minute': 0,
        'completion_time': ''
    }

    for line in lines:
        if 'Всего тестов:' in line:
            report_data['accuracy']['total_tests'] = int(line.split(':')[-1].strip())
        elif 'Показания: верно' in line:
            parts = line.split('верно')[-1].split('/')
            report_data['accuracy']['indications']['correct'] = int(parts[0].strip())
            report_data['accuracy']['indications']['accuracy'] = float(
                parts[1].split('(')[-1].replace('%)', '').strip())
        elif 'Серийные номера: верно' in line:
            parts = line.split('верно')[-1].split('/')
            report_data['accuracy']['series']['correct'] = int(parts[0].strip())
            report_data['accuracy']['series']['accuracy'] = float(parts[1].split('(')[-1].replace('%)', '').strip())
        elif 'Модели: верно' in line:
            parts = line.split('верно')[-1].split('/')
            report_data['accuracy']['model']['correct'] = int(parts[0].strip())
            report_data['accuracy']['model']['accuracy'] = float(parts[1].split('(')[-1].replace('%)', '').strip())
        elif 'Тарифы: верно' in line:
            parts = line.split('верно')[-1].split('/')
            report_data['accuracy']['rate']['correct'] = int(parts[0].strip())
            report_data['accuracy']['rate']['accuracy'] = float(parts[1].split('(')[-1].replace('%)', '').strip())
        elif 'Общая точность:' in line:
            report_data['accuracy']['overall']['accuracy'] = float(line.split(':')[-1].replace('%', '').strip())
        elif 'Всего файлов:' in line:
            report_data['total_images'] = int(line.split(':')[-1].strip())
        elif 'Успешно обработано:' in line:
            report_data['successfully_processed'] = int(line.split(':')[-1].strip())
        elif 'Ошибок:' in line:
            report_data['errors'] = int(line.split(':')[-1].strip())
        elif 'Пропущено:' in line:
            report_data['skipped'] = int(line.split(':')[-1].strip())
        elif 'Успешность обработки:' in line:
            report_data['success_rate'] = float(line.split(':')[-1].replace('%', '').strip())
        elif 'Общее время:' in line:
            report_data['total_time_seconds'] = float(line.split(':')[-1].replace('секунд', '').strip())
        elif 'Среднее время на изображение:' in line:
            report_data['average_time_per_image'] = float(line.split(':')[-1].replace('секунд', '').strip())
        elif 'Скорость обработки:' in line:
            report_data['images_per_minute'] = float(line.split(':')[-1].replace('изображений/мин', '').strip())
        elif 'Завершено:' in line:
            report_data['completion_time'] = line.split('Завершено:')[-1].strip()

    return report_data