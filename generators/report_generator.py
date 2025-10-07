import logging
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from config import *
from accuracy_calculator import calculate_accuracy_stats

logger = logging.getLogger(__name__)


def add_version_info_to_excel(wb, excel_file_path):
    try:
        ws = wb['Image Data']
        version = get_git_version()
        creation_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws.insert_rows(1, 3)
        ws.merge_cells('A1:B1')
        version_cell = ws.cell(row=1, column=1, value=f"Версия приложения: v{version}")
        version_cell.font = Font(bold=True, size=12)
        version_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('C1:D1')
        time_cell = ws.cell(row=1, column=3, value=f"Время тестирования: {creation_time}")
        time_cell.font = Font(bold=True, size=12)
        time_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A2:D2')
        file_cell = ws.cell(row=2, column=1, value=f"Файл результатов: {os.path.basename(excel_file_path)}")
        file_cell.font = Font(italic=True, size=10)
        file_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws.merge_cells('A3:D3')
        separator_cell = ws.cell(row=3, column=1, value="=" * 80)
        separator_cell.alignment = Alignment(horizontal='center', vertical='center')
        separator_cell.font = Font(size=8, color="808080")

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 25

        logger.info(f"Добавлена информация о версии v{version} и времени {creation_time} в файл")

    except Exception as e:
        logger.warning(f"Не удалось добавить информацию о версии в Excel: {e}")


def apply_excel_styles(wb, excel_file_path):
    ws = wb['Image Data']
    add_version_info_to_excel(wb, excel_file_path)

    headers = [cell.value for cell in ws[4]]
    if 'Overall Confidence' not in headers:
        try:
            confidences_idx = headers.index('Recognition Confidence') + 1
            ws.insert_cols(confidences_idx + 1)
            ws.insert_cols(row=4, column=confidences_idx + 1, value='Overall Confidence')
        except:
            ws.cell(row=4, column=ws.max_column + 1, value='Overall Confidence')

    data_start_row = 4

    reference_columns = []
    for col_idx, col_name in enumerate(ws[data_start_row], 1):
        if col_name.value and 'reference' in str(col_name.value).lower():
            reference_columns.append(col_idx)
    for row in ws.iter_rows(min_row=data_start_row + 1):
        for cell in row:
            if cell.column in [13, 14, 15, 16, 17]:
                cell.number_format = '0'
            else:
                cell.number_format = '@'

            cell.border = Border(left=THIN_BORDER, right=THIN_BORDER,
                                 top=THIN_BORDER, bottom=THIN_BORDER)
            cell.alignment = CENTER_ALIGNMENT

    REFERENCE_FILL = PatternFill(start_color="E6E6FA", end_color="E6E6FA",
                                 fill_type="solid")
    RESULT_FILL = PatternFill(start_color="E0FFFF", end_color="E0FFFF",
                              fill_type="solid")
    FILENAME_FILL = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    DIMENSIONS_FILL = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")
    MISMATCH_FILL = PatternFill(start_color="FFCCCB", end_color="FFCCCB", fill_type="solid")

    for col in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=data_start_row, column=col)
        header_cell.fill = HEADER_FILL
        header_cell.font = BOLD_FONT

    for row in range(data_start_row + 1, ws.max_row + 1):
        ws.cell(row=row, column=1).fill = FILENAME_FILL

        for col in range(2, 5):
            ws.cell(row=row, column=col).fill = DIMENSIONS_FILL

        for col in range(5, 9):
            ws.cell(row=row, column=col).fill = REFERENCE_FILL

        apply_data_mismatch_highlighting(ws, row, MISMATCH_FILL, RESULT_FILL)

    apply_match_column_styles(ws, data_start_row)
    apply_table_border(ws, data_start_row)


def apply_data_mismatch_highlighting(ws, row, mismatch_fill, result_fill):
    column_mapping = {
        9: 5,
        10: 6,
        11: 7,
        12: 8
    }

    for result_col, ref_col in column_mapping.items():
        result_cell = ws.cell(row=row, column=result_col)
        ref_cell = ws.cell(row=row, column=ref_col)

        result_value = str(result_cell.value).strip().lower() if result_cell.value is not None else ""
        ref_value = str(ref_cell.value).strip().lower() if ref_cell.value is not None else ""

        if result_value != ref_value:
            result_cell.fill = mismatch_fill
        else:
            result_cell.fill = result_fill


def apply_match_column_styles(ws, start_row=4):
    for row in range(start_row + 1, ws.max_row + 1):
        for col in range(13, 18):
            cell = ws.cell(row=row, column=col)
            if cell.value == 1:
                cell.fill = GREEN_FILL
            elif cell.value == 0:
                cell.fill = RED_FILL


def apply_table_border(ws, start_row=4):
    first_row, last_row = start_row, ws.max_row
    first_col, last_col = 1, ws.max_column

    for row in range(first_row, last_row + 1):
        for col in [first_col, last_col]:
            cell = ws.cell(row=row, column=col)
            current_border = cell.border
            new_border = Border(
                left=THICK_BORDER if col == first_col else current_border.left,
                right=THICK_BORDER if col == last_col else current_border.right,
                top=THICK_BORDER if row == first_row else current_border.top,
                bottom=THICK_BORDER if row == last_row else current_border.bottom
            )
            cell.border = new_border


def auto_adjust_column_widths(ws, start_row=4):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            if cell.row >= start_row:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass

        adjusted_width = min((max_length + 2), 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_summary_report(processed_count, errors_count, skipped_count, total_time, excel_file):
    logger.info(f"🎯 ПОЛУЧЕН ФАЙЛ В generate_summary_report: {excel_file}")
    logger.info(f"📁 Абсолютный путь: {os.path.abspath(excel_file)}")

    total_attempted = processed_count + errors_count
    timing_totals = []

    try:
        logger.info(f"📊 ЗАГРУЖАЕМ ДАННЫЕ ДЛЯ РАСЧЕТА ТОЧНОСТИ ИЗ: {excel_file}")

        df_full = pd.read_excel(excel_file, sheet_name='Image Data', header=None)
        logger.info(f"📋 ВСЕГО СТРОК В ФАЙЛЕ: {len(df_full)}")

        header_row = None
        for i in range(min(10, len(df_full))):
            row_values = [str(x) for x in df_full.iloc[i].values if pd.notna(x)]
            logger.info(f"   Строка {i}: {row_values}")
            if 'Filename' in row_values:
                header_row = i
                logger.info(f"✅ НАЙДЕНЫ ЗАГОЛОВКИ В СТРОКЕ {i}")
                break

        if header_row is None:
            logger.error("❌ НЕ НАЙДЕНА СТРОКА С ЗАГОЛОВКАМИ")
            accuracy_stats = create_empty_accuracy_stats()
        else:
            df = pd.read_excel(excel_file, sheet_name='Image Data', header=header_row)
            logger.info(f"📊 ЗАГРУЖЕНО ДАННЫХ: {len(df)} строк, {len(df.columns)} колонок")
            logger.info(f"📋 КОЛОНКИ: {list(df.columns)}")

            # ПОЛУЧАЕМ TIMING TOTALS ИЗ DataFrame
            if 'Timing Total' in df.columns:
                timing_totals = df['Timing Total'].dropna().tolist()
                logger.info(f"⏱️  Найдено {len(timing_totals)} значений Timing Total")
                if timing_totals:
                    avg_from_timing = sum(timing_totals) / len(timing_totals)
                    logger.info(f"📊 Среднее время из Timing Total: {avg_from_timing:.2f} сек")
            else:
                logger.warning("⚠️ Колонка 'Timing Total' не найдена в DataFrame")

            required_cols = ['Indications Match', 'Series Match', 'Model Match', 'Rate Match', 'Overall Match']
            for col in required_cols:
                if col in df.columns:
                    non_zero = (df[col] == 1).sum()
                    logger.info(f"   ✅ {col}: {non_zero} совпадений из {len(df)}")
                else:
                    logger.error(f"   ❌ {col}: ОТСУТСТВУЕТ")

            accuracy_stats = calculate_accuracy_stats(df)

    except Exception as e:
        logger.error(f"❌ Ошибка загрузки данных для расчета точности: {str(e)}")
        import traceback
        logger.error(f"📋 Детали ошибки: {traceback.format_exc()}")
        accuracy_stats = create_empty_accuracy_stats()

    report = create_report_dict(processed_count, errors_count, skipped_count,
                                total_attempted, total_time, accuracy_stats, timing_totals)

    print_report(report)

    try:
        wb = openpyxl.load_workbook(excel_file)
        from generators.summary_report import create_summary_sheet
        create_summary_sheet(wb, report)
        wb.save(excel_file)
        logger.info("✅ Итоговый отчет добавлен в Excel файл")
    except Exception as e:
        logger.error(f"❌ Ошибка сохранения отчета в Excel: {str(e)}")
        import traceback
        logger.error(f"📋 Детали ошибки: {traceback.format_exc()}")

    return report


def create_empty_accuracy_stats():
    return {
        'total_tests': 0,
        'indications': {'correct': 0, 'accuracy': 0},
        'series': {'correct': 0, 'accuracy': 0},
        'model': {'correct': 0, 'accuracy': 0},
        'rate': {'correct': 0, 'accuracy': 0},
        'overall': {'correct': 0, 'accuracy': 0}
    }


# def generate_summary_report(processed_count, errors_count, skipped_count, total_time, excel_file):
#     logger.info(f"🎯 ПОЛУЧЕН ФАЙЛ В generate_summary_report: {excel_file}")
#     logger.info(f"📁 Абсолютный путь: {os.path.abspath(excel_file)}")
#
#     total_attempted = processed_count + errors_count
#     timing_totals = []
#
#     try:
#         logger.info(f"📊 ЗАГРУЖАЕМ ДАННЫЕ ДЛЯ РАСЧЕТА ТОЧНОСТИ ИЗ: {excel_file}")
#
#         df_full = pd.read_excel(excel_file, sheet_name='Image Data', header=None)
#         logger.info(f"📋 ВСЕГО СТРОК В ФАЙЛЕ: {len(df_full)}")
#
#         header_row = None
#         for i in range(min(10, len(df_full))):
#             row_values = [str(x) for x in df_full.iloc[i].values if pd.notna(x)]
#             logger.info(f"   Строка {i}: {row_values}")
#             if 'Filename' in row_values:
#                 header_row = i
#                 logger.info(f"✅ НАЙДЕНЫ ЗАГОЛОВКИ В СТРОКЕ {i}")
#                 break
#
#         if header_row is None:
#             logger.error("❌ НЕ НАЙДЕНА СТРОКА С ЗАГОЛОВКАМИ")
#             accuracy_stats = create_empty_accuracy_stats()
#         else:
#             df = pd.read_excel(excel_file, sheet_name='Image Data', header=header_row)
#             logger.info(f"📊 ЗАГРУЖЕНО ДАННЫХ: {len(df)} строк, {len(df.columns)} колонок")
#             logger.info(f"📋 КОЛОНКИ: {list(df.columns)}")
#
#             # ПОЛУЧАЕМ TIMING TOTALS ИЗ DataFrame
#             if 'Timing Total' in df.columns:
#                 timing_totals = df['Timing Total'].dropna().tolist()
#                 logger.info(f"⏱️  Найдено {len(timing_totals)} значений Timing Total")
#                 if timing_totals:
#                     avg_from_timing = sum(timing_totals) / len(timing_totals)
#                     logger.info(f"📊 Среднее время из Timing Total: {avg_from_timing:.2f} сек")
#             else:
#                 logger.warning("⚠️ Колонка 'Timing Total' не найдена в DataFrame")
#
#             required_cols = ['Indications Match', 'Series Match', 'Model Match', 'Rate Match', 'Overall Match']
#             for col in required_cols:
#                 if col in df.columns:
#                     non_zero = (df[col] == 1).sum()
#                     logger.info(f"   ✅ {col}: {non_zero} совпадений из {len(df)}")
#                 else:
#                     logger.error(f"   ❌ {col}: ОТСУТСТВУЕТ")
#
#             accuracy_stats = calculate_accuracy_stats(df)
#
#     except Exception as e:
#         logger.error(f"❌ Ошибка загрузки данных для расчета точности: {str(e)}")
#         import traceback
#         logger.error(f"📋 Детали ошибки: {traceback.format_exc()}")
#         accuracy_stats = create_empty_accuracy_stats()
#
#     # Передаем timing_totals в создание отчета
#     report = create_report_dict(processed_count, errors_count, skipped_count,
#                                 total_attempted, total_time, accuracy_stats, timing_totals)
#
#     print_report(report)
#
#     try:
#         wb = openpyxl.load_workbook(excel_file)
#         from generators.summary_report import create_summary_sheet
#         create_summary_sheet(wb, report)
#         wb.save(excel_file)
#         logger.info("✅ Итоговый отчет добавлен в Excel файл")
#     except Exception as e:
#         logger.error(f"❌ Ошибка сохранения отчета в Excel: {str(e)}")
#         import traceback
#         logger.error(f"📋 Детали ошибки: {traceback.format_exc()}")
#
#     return report


def create_report_dict(processed, errors, skipped, attempted, total_time, accuracy_stats, timing_totals=None):
    if timing_totals and len(timing_totals) > 0:
        average_time = sum(timing_totals) / len(timing_totals)
        logger.info(f"✅ Среднее время рассчитано из Timing Total: {average_time:.2f} сек")
    else:
        average_time = total_time / attempted if attempted > 0 else 0
        logger.info(f"⚠️ Среднее время рассчитано из общего времени: {average_time:.2f} сек")

    return {
        "total_images": attempted + skipped,
        "successfully_processed": processed,
        "errors": errors,
        "skipped": skipped,
        "success_rate": (processed / attempted) * 100 if attempted > 0 else 0,
        "total_time_seconds": total_time,
        "average_time_per_image": average_time,
        "images_per_minute": (attempted / total_time) * 60 if total_time > 0 else 0,
        "completion_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "accuracy": accuracy_stats
    }


def print_report(report):
    acc = report['accuracy']

    logger.info("=" * 60)
    logger.info("ИТОГОВЫЙ ОТЧЕТ С ТОЧНОСТЬЮ")
    logger.info("=" * 60)
    logger.info(f"Всего тестов: {acc['total_tests']}")
    logger.info("=" * 50)
    logger.info(
        f"Показания: верно {acc['indications']['correct']}/{acc['total_tests']} ({acc['indications']['accuracy']:.1f}%)")
    logger.info(
        f"Серийные номера: верно {acc['series']['correct']}/{acc['total_tests']} ({acc['series']['accuracy']:.1f}%)")
    logger.info(f"Модели: верно {acc['model']['correct']}/{acc['total_tests']} ({acc['model']['accuracy']:.1f}%)")
    logger.info(f"Тарифы: верно {acc['rate']['correct']}/{acc['total_tests']} ({acc['rate']['accuracy']:.1f}%)")
    logger.info("=" * 50)
    logger.info(f"Общая точность: {acc['overall']['accuracy']:.1f}%")
    logger.info("=" * 50)
    logger.info(f"Всего файлов: {report['total_images']}")
    logger.info(f"Успешно обработано: {report['successfully_processed']}")
    logger.info(f"Ошибок: {report['errors']}")
    logger.info(f"Пропущено: {report['skipped']}")
    logger.info(f"Успешность обработки: {report['success_rate']:.2f}%")
    logger.info(f"Общее время: {report['total_time_seconds']:.2f} секунд")
    logger.info(f"Среднее время на изображение: {report['average_time_per_image']:.2f} секунд")
    logger.info(f"Скорость обработки: {report['images_per_minute']:.2f} изображений/мин")
    logger.info(f"Завершено: {report['completion_time']}")
    logger.info("=" * 60)