import threading
from datetime import datetime
from dotenv import load_dotenv
from openpyxl.styles import PatternFill, Alignment, Font, Border, Side
import subprocess
import os
import logging

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
)
logger = logging.getLogger(__name__)

load_dotenv()

# ОТЛАДОЧНАЯ ИНФОРМАЦИЯ ДО ПРЕОБРАЗОВАНИЯ
logger.info("🔧 ЗАГРУЖЕННЫЕ ПЕРЕМЕННЫЕ ИЗ .env:")
logger.info(f"   MAX_WORKERS (raw): '{os.getenv('MAX_WORKERS')}'")
logger.info(f"   PROCESSING_MODE: '{os.getenv('PROCESSING_MODE')}'")
logger.info(f"   SELECTED_SERVER: '{os.getenv('SELECTED_SERVER')}'")

MAIN_REPO_PATH = os.getenv('MAIN_REPO_PATH', '/app')
FOLDER_TEST = os.getenv('FOLDER_TEST', '/app/testing_sets/test_1')
EXCEL_DATA = os.getenv('EXCEL_DATA', '/app/demo/Тестирование_1.xlsx')

PROGRAM_SCRIPT = os.getenv('PROGRAM_SCRIPT', '')

SELECTED_SERVER = os.getenv('SELECTED_SERVER', 'server2')
AUTHORIZED_TOKEN = os.getenv('AUTHORIZED_TOKEN', '8DWQLfproEJlyC8dJaLqRhBx1B2sJyZR4V')

DB_HOST = os.getenv('DB_HOST', 'mysql')
DB_PORT = int(os.getenv('DB_PORT', 3306))
DB_NAME = os.getenv('DB_NAME', 'testing_system')
DB_USER = os.getenv('DB_USER', 'root')
DB_PASSWORD = os.getenv('DB_PASSWORD', 'root')

SERVERS = {
    'default': 'default',
    'server1': 'http://80.93.179.130:9099',
    'server2': 'http://ai-server1.ugrey.ru/'
}

PROCESSING_MODE = os.getenv('PROCESSING_MODE', 'sequential')

# ИСПРАВЛЕННОЕ ПРЕОБРАЗОВАНИЕ MAX_WORKERS
max_workers_str = os.getenv('MAX_WORKERS', '1')
try:
    MAX_WORKERS = int(max_workers_str)
    logger.info(f"✅ MAX_WORKERS преобразован в int: {MAX_WORKERS}")
except (ValueError, TypeError) as e:
    logger.error(f"❌ Ошибка преобразования MAX_WORKERS '{max_workers_str}': {e}")
    MAX_WORKERS = 1

# ОТЛАДОЧНАЯ ИНФОРМАЦИЯ ПОСЛЕ ПРЕОБРАЗОВАНИЯ
logger.info("🔧 ФИНАЛЬНЫЕ ЗНАЧЕНИЯ КОНФИГУРАЦИИ:")
logger.info(f"   MAX_WORKERS: {MAX_WORKERS} (тип: {type(MAX_WORKERS)})")
logger.info(f"   PROCESSING_MODE: '{PROCESSING_MODE}'")
logger.info(f"   SELECTED_SERVER: '{SELECTED_SERVER}'")

def get_git_version():
    try:
        result = subprocess.run(
            ['git', 'describe', '--tags', '--abbrev=0'],
            capture_output=True,
            text=True,
            cwd=MAIN_REPO_PATH
        )
        if result.returncode == 0:
            return result.stdout.strip().replace('v', '')

        result = subprocess.run(
            ['git', 'rev-parse', '--short', 'HEAD'],
            capture_output=True,
            text=True,
            cwd=MAIN_REPO_PATH
        )
        if result.returncode == 0:
            return f"dev-{result.stdout.strip()}"

        return "unknown"
    except Exception as e:
        logger.warning(f"Не удалось получить версию из git: {e}")
        return "unknown"

def rename_file_with_version_and_time(original_path):
    if not os.path.exists(original_path):
        logger.warning(f"Файл для переименования не существует: {original_path}")
        return original_path

    try:
        directory = os.path.dirname(original_path)
        filename = os.path.basename(original_path)
        name, ext = os.path.splitext(filename)

        version = get_git_version()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        new_filename = f"{name}_v{version}_{timestamp}{ext}"
        new_path = os.path.join(directory, new_filename)

        os.rename(original_path, new_path)
        logger.info(f"Файл переименован: {new_path}")
        return new_path
    except Exception as e:
        logger.error(f"Ошибка переименования файла {original_path}: {e}")
        return original_path

APP_VERSION = get_git_version()

processed_count = 0
errors_count = 0
skipped_count = 0
df_lock = threading.Lock()

GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
BLUE_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')
BOLD_FONT = Font(bold=True)

THIN_BORDER = Side(border_style="thin", color="000000")
THICK_BORDER = Side(border_style="thick", color="000000")
HEADER_FILL = BLUE_FILL

TIMEOUT = 120
SUPPORTED_IMAGE_EXTENSIONS = ('.jpg', '.jpeg', '.png', '.bmp', '.tiff', '.tif')

logger.info("=" * 60)
logger.info("⚙️  КОНФИГУРАЦИЯ АВТОТЕСТОВ")
logger.info("=" * 60)
logger.info(f"📁 MAIN_REPO_PATH: {MAIN_REPO_PATH}")
logger.info(f"📊 EXCEL_DATA: {EXCEL_DATA}")
logger.info(f"🖼️  FOLDER_TEST: {FOLDER_TEST}")
logger.info(f"🌐 SELECTED_SERVER: {SELECTED_SERVER}")
logger.info(f"🗄️  MySQL: {DB_HOST}:{DB_PORT}/{DB_NAME}")
logger.info(f"🔐 TOKEN: {AUTHORIZED_TOKEN[:8]}...")
logger.info(f"🏷️  VERSION: {APP_VERSION}")
logger.info("=" * 60)

def check_required_files():
    errors = []

    if not os.path.exists(EXCEL_DATA):
        errors.append(f"Excel файл не найден: {EXCEL_DATA}")

    if not os.path.exists(FOLDER_TEST):
        errors.append(f"Папка с изображениями не найдена: {FOLDER_TEST}")
    else:
        image_files = [f for f in os.listdir(FOLDER_TEST)
                       if f.lower().endswith(SUPPORTED_IMAGE_EXTENSIONS)]
        if not image_files:
            errors.append(f"В папке {FOLDER_TEST} нет поддерживаемых изображений")
        else:
            logger.info(f"📸 Найдено изображений: {len(image_files)}")

    if errors:
        for error in errors:
            logger.error(f"❌ {error}")
        return False

    logger.info("✅ Все необходимые файлы найдены")
    return True

if __name__ != "__main__":
    check_required_files()